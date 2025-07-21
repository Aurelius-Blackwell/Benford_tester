import openpyxl
from openpyxl.utils import get_column_letter
import re
import os

# The percentage frequency of first digits will be written to this Excel template
doc = 'benfordTests_template.xlsx'
wb = openpyxl.load_workbook(doc)
sheet = wb['sheetname']

# Folder of files to be analyzed
folder = 'C:/your_folder_here/'

# Place folder of files whose contents will be read
p = os.listdir(folder)

# Start at row for to leave some commenting space in the Excel template
row = 4
for file in p:
    # The A-column holds the name of the document being read
    name = file.rstrip('.txt')
    sheet['a'+str(row)].value = name

    # Progress indicator
    print('Working on ' + name)

    # Open and read file
    f = open(folder+file, 'r', encoding='utf-8')
    text = f.read()

    # List all indepedent collect of digits
    numList = re.findall('\d+', text)

    # Totals of each digit, resets for each file
    totals = [0,0,0,0,0,0,0,0,0]            
    # Count first digits
    for i in range(len(numList)):
        x = int(numList[i])
        firstDigit = int(str(x)[0])
        totals[firstDigit-1] +=1    #-1 to place digit in totals by index logic
    # Calculate proportion of first digits
    for i in range(len(totals)):
        totals[i] = totals[i]/len(numList)
        sheet[get_column_letter(i+3)+str(row)].value = totals[i]    #+3 to start at column C
    # Calculate totals in L-column to check that distribution totals 100%
    sheet['l'+str(row)] = '=sum(c'+str(row)+':k'+str(row)+')'
    wb.save('benfordTests.xlsx')
    # Write next file's distribution onto new row
    row +=1

# Close workbook at end of program to prevent bugs
wb.close()
